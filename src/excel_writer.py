"""
Excel Writer - Outputs job data to Excel file

Creates a nicely formatted Excel spreadsheet with all job listings.
"""

from pathlib import Path
from typing import List, Union
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from src.config import config
from src.linkedin_scraper import JobListing
from src.ai_filter import FilteredJob


class ExcelWriter:
    """Writes job data to Excel file with nice formatting."""
    
    # Column configuration
    COLUMNS = [
        ("Title", 40),
        ("Company", 25),
        ("Location", 20),
        ("Type", 15),
        ("Posted", 12),
        ("Score", 8),
        ("AI Analysis", 50),
        ("Job URL", 60),
        ("Scraped At", 20),
    ]
    
    # Styles
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    GOOD_SCORE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    AVG_SCORE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    def __init__(self, output_path: Union[str, Path] = None):
        self.output_path = Path(output_path) if output_path else config.get_output_path()
        self.output_path.parent.mkdir(exist_ok=True)
    
    def write(self, jobs: List[Union[JobListing, FilteredJob]]) -> Path:
        """
        Write jobs to Excel file.
        
        Args:
            jobs: List of JobListing or FilteredJob objects
            
        Returns:
            Path to the created Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Freelance Gigs"
        
        # Write header
        self._write_header(ws)
        
        # Write data rows
        for i, job in enumerate(jobs, start=2):
            self._write_row(ws, i, job)
        
        # Auto-adjust column widths
        self._adjust_columns(ws)
        
        # Add filters
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUMNS))}1"
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save
        wb.save(self.output_path)
        logger.info(f"Saved {len(jobs)} jobs to: {self.output_path}")
        
        return self.output_path
    
    def _write_header(self, ws):
        """Write the header row."""
        for col, (name, _) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _write_row(self, ws, row: int, job: Union[JobListing, FilteredJob]):
        """Write a single data row."""
        
        # Handle both JobListing and FilteredJob
        if isinstance(job, FilteredJob):
            job_data = job.job
            score = job.quality_score
            analysis = job.reason
        else:
            job_data = job
            score = "N/A"
            analysis = "Not analyzed"
        
        values = [
            job_data.title,
            job_data.company,
            job_data.location,
            job_data.job_type,
            job_data.posted_date,
            score,
            analysis,
            job_data.job_url,
            job_data.scraped_at,
        ]
        
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Color code scores
            if col == 6 and isinstance(score, int):  # Score column
                if score >= 7:
                    cell.fill = self.GOOD_SCORE_FILL
                elif score >= 5:
                    cell.fill = self.AVG_SCORE_FILL
            
            # Make URL clickable
            if col == 8 and value and value != "N/A":
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
    
    def _adjust_columns(self, ws):
        """Set column widths."""
        for col, (_, width) in enumerate(self.COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width


def write_to_excel(
    jobs: List[Union[JobListing, FilteredJob]],
    output_path: Union[str, Path] = None
) -> Path:
    """
    Convenience function to write jobs to Excel.
    
    Args:
        jobs: List of JobListing or FilteredJob objects
        output_path: Optional custom output path
        
    Returns:
        Path to the created Excel file
    """
    writer = ExcelWriter(output_path)
    return writer.write(jobs)


if __name__ == "__main__":
    # Test with sample data
    sample_jobs = [
        JobListing(
            title="Python Developer - Contract",
            company="TechCorp",
            location="Remote, USA",
            job_type="Contract",
            posted_date="2024-01-15",
            description="Build APIs",
            job_url="https://linkedin.com/jobs/123"
        ),
        JobListing(
            title="FastAPI Expert",
            company="StartupX",
            location="New York, NY",
            job_type="Freelance",
            posted_date="2024-01-14",
            description="Backend work",
            job_url="https://linkedin.com/jobs/456"
        )
    ]
    
    output = write_to_excel(sample_jobs, "test_output.xlsx")
    print(f"Created: {output}")
